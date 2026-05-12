#!/usr/bin/env python3
"""
Build a first-pass PTE knowledge draft from a local course resource folder.

This script is intentionally conservative:
- It reads source files from the mounted course library.
- It skips video/audio archives for v1.
- It extracts bounded text from PDF/DOCX/XLSX/TXT/MD only.
- It writes reviewable draft JSON and reports.
- It does not modify the live website knowledge base.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any


SUPPORTED_EXTS = {".pdf", ".docx", ".xlsx", ".txt", ".md"}
MEDIA_EXTS = {".mp4", ".mov", ".m4v", ".mp3", ".m4a", ".wav", ".aac"}
ARCHIVE_EXTS = {".zip", ".rar", ".7z"}

TAG_RULES = [
    ("RA", ["ra", "read aloud", "朗读"]),
    ("RS", ["rs", "repeat sentence", "复述"]),
    ("DI", ["di", "describe image", "图表", "描述图片"]),
    ("RL", ["rl", "retell lecture", "复述讲座"]),
    ("ASQ", ["asq", "answer short question"]),
    ("SWT", ["swt", "summarize written text", "小作文", "总结书面文本"]),
    ("WE", ["we", "essay", "大作文", "写作"]),
    ("WFD", ["wfd", "write from dictation", "听写"]),
    ("SST", ["sst", "summarize spoken text"]),
    ("FIB", ["fib", "fill in", "完形", "填空"]),
    ("RO", ["ro", "re-order", "排序"]),
    ("HIW", ["hiw", "highlight incorrect words"]),
    ("PTE Core", ["pte core", "core"]),
    ("口语", ["口语", "speaking"]),
    ("听力", ["听力", "listening"]),
    ("阅读", ["阅读", "reading"]),
    ("写作", ["写作", "writing"]),
    ("词汇", ["词汇", "vocab", "vocabulary", "单词"]),
    ("语法", ["语法", "grammar"]),
    ("模板", ["模板", "模版", "template", "框架"]),
    ("机经", ["机经", "真题", "预测", "高频"]),
    ("学习计划", ["学习计划", "备考计划", "study plan"]),
    ("报考流程", ["报考", "考试流程", "考场", "报名"]),
    ("评分", ["评分", "score guide", "分数"]),
]

HIGH_PRIORITY = [
    "2026",
    "改革后",
    "优先",
    "技巧直播课",
    "技巧真经班",
    "备考资料包",
    "最新模板",
    "学习计划",
    "官方",
    "score guide",
    "入门",
]

LOW_PRIORITY = ["历史", "18-23", "2024年技巧课程汇总", "旧", "改革前"]


def clean_text(text: str, max_chars: int = 8000) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = text.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore")
    return text.strip()[:max_chars]


def sanitize_for_json(value: Any) -> Any:
    if isinstance(value, str):
        return value.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore")
    if isinstance(value, list):
        return [sanitize_for_json(item) for item in value]
    if isinstance(value, dict):
        return {sanitize_for_json(k): sanitize_for_json(v) for k, v in value.items()}
    return value


def safe_id(path: Path) -> str:
    digest = hashlib.sha1(str(path).encode("utf-8")).hexdigest()[:10]
    stem = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", path.stem).strip("-")
    return f"{stem[:36]}-{digest}" if stem else digest


def detect_tags(path: Path, text: str = "") -> list[str]:
    haystack = f"{path} {text[:3000]}".lower()
    tags = []
    for tag, needles in TAG_RULES:
        if any(needle.lower() in haystack for needle in needles):
            tags.append(tag)
    return tags


def score_priority(path: Path, tags: list[str], size_mb: float) -> int:
    haystack = str(path).lower()
    score = 0
    for key in HIGH_PRIORITY:
        if key.lower() in haystack:
            score += 4
    for key in LOW_PRIORITY:
        if key.lower() in haystack:
            score -= 3
    score += min(len(tags), 8)
    if size_mb > 80:
        score -= 2
    return score


def make_questions(tags: list[str], title: str) -> list[str]:
    questions = []
    if "学习计划" in tags:
        questions.append("PTE 应该怎么安排备考计划？")
    if "RA" in tags or "口语" in tags:
        questions.append("PTE 口语 RA 怎么提分？")
    if "DI" in tags or "RL" in tags:
        questions.append("DI 和 RL 模板应该怎么使用？")
    if "WE" in tags or "SWT" in tags or "写作" in tags:
        questions.append("PTE 写作模板怎么练才稳？")
    if "WFD" in tags or "SST" in tags or "听力" in tags:
        questions.append("PTE 听力 WFD 和 SST 怎么训练？")
    if "FIB" in tags or "RO" in tags or "阅读" in tags:
        questions.append("PTE 阅读 FIB 和 RO 怎么练？")
    if "报考流程" in tags:
        questions.append("PTE 报名和考试流程要注意什么？")
    if "评分" in tags:
        questions.append("PTE 评分规则应该怎么看？")
    if not questions:
        questions.append(f"{title} 这份资料适合解决什么问题？")
    return questions[:5]


def answer_draft(title: str, tags: list[str], rel_path: str) -> str:
    tag_text = "、".join(tags[:6]) if tags else "PTE 备考"
    return (
        f"这条知识来自课程资料《{title}》，主题与{tag_text}相关。"
        "当前为夜间自动生成的知识库草稿，建议顾问审核后再用于 Sunny 正式回答。"
        f"来源路径：{rel_path}"
    )


def extract_pdf(path: Path, max_pages: int, max_chars: int) -> tuple[str, str]:
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        parts = []
        for page in reader.pages[:max_pages]:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                continue
        text = clean_text("\n".join(parts), max_chars)
        return text, "ok" if text else "empty_pdf_text"
    except Exception as exc:
        return "", f"pdf_error:{type(exc).__name__}"


def extract_docx(path: Path, max_chars: int) -> tuple[str, str]:
    try:
        import docx

        doc = docx.Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return clean_text(text, max_chars), "ok" if text.strip() else "empty_docx_text"
    except Exception as exc:
        return "", f"docx_error:{type(exc).__name__}"


def extract_xlsx(path: Path, max_chars: int, max_rows: int = 80) -> tuple[str, str]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames[:4]:
            ws = wb[sheet_name]
            parts.append(f"Sheet: {sheet_name}")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if values:
                    parts.append(" | ".join(values[:8]))
                    row_count += 1
                if row_count >= max_rows:
                    break
        text = clean_text("\n".join(parts), max_chars)
        return text, "ok" if text else "empty_xlsx_text"
    except Exception as exc:
        return "", f"xlsx_error:{type(exc).__name__}"


def extract_plain(path: Path, max_chars: int) -> tuple[str, str]:
    for encoding in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return clean_text(path.read_text(encoding=encoding, errors="ignore"), max_chars), "ok"
        except Exception:
            continue
    return "", "plain_error"


def extract_text(path: Path, max_pages: int, max_chars: int) -> tuple[str, str]:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf(path, max_pages, max_chars)
    if ext == ".docx":
        return extract_docx(path, max_chars)
    if ext == ".xlsx":
        return extract_xlsx(path, max_chars)
    if ext in {".txt", ".md"}:
        return extract_plain(path, max_chars)
    return "", "unsupported"


def iter_files(root: Path) -> list[Path]:
    return [p for p in root.rglob("*") if p.is_file() and p.name != ".DS_Store"]


def build(args: argparse.Namespace) -> dict[str, Any]:
    root = Path(args.source_root).expanduser().resolve()
    if not root.exists():
        raise SystemExit(f"Source root does not exist: {root}")

    run_id = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_dir = Path(args.output_dir).expanduser().resolve() / run_id
    out_dir.mkdir(parents=True, exist_ok=True)

    all_files = iter_files(root)
    ext_counter = Counter(p.suffix.lower() or "[no extension]" for p in all_files)
    candidates = [p for p in all_files if p.suffix.lower() in SUPPORTED_EXTS]

    if args.max_files:
        candidates = candidates[: args.max_files]

    manifest_rows = []
    review_entries = []
    sunny_entries = []
    status_counter = Counter()

    for i, path in enumerate(candidates, start=1):
        try:
            size = path.stat().st_size
            size_mb = round(size / 1024 / 1024, 2)
            rel = str(path.relative_to(root))
        except Exception:
            continue

        if size_mb > args.max_extract_mb:
            text, status = "", "skipped_large_file"
        else:
            text, status = extract_text(path, args.max_pdf_pages, args.max_chars)

        tags = detect_tags(path, text)
        priority = score_priority(path, tags, size_mb)
        title = path.stem
        entry_id = safe_id(path)
        questions = make_questions(tags, title)

        status_counter[status] += 1
        manifest_rows.append({
            "path": str(path),
            "relative_path": rel,
            "extension": path.suffix.lower(),
            "size_mb": size_mb,
            "status": status,
            "priority": priority,
            "tags": ";".join(tags),
        })

        review_entries.append({
            "id": entry_id,
            "title": title,
            "relativePath": rel,
            "sourcePath": str(path),
            "fileType": path.suffix.lower().lstrip("."),
            "sizeMB": size_mb,
            "priority": priority,
            "tags": tags,
            "suggestedQuestions": questions,
            "extractionStatus": status,
            "textPreview": text[:1200],
            "needsHumanReview": True,
        })

        sunny_entries.append({
            "id": entry_id,
            "title": title,
            "keywords": list(dict.fromkeys(tags + [title] + questions))[:18],
            "answer": answer_draft(title, tags, rel),
            "source": rel,
            "status": "draft_needs_review",
        })

        if i % 25 == 0:
            print(f"Processed {i}/{len(candidates)} document candidates...", flush=True)

    review_entries.sort(key=lambda e: (e["priority"], -e["sizeMB"]), reverse=True)
    sunny_entries_by_id = {e["id"]: e for e in sunny_entries}
    sunny_entries = [sunny_entries_by_id[e["id"]] for e in review_entries]

    summary = {
        "runId": run_id,
        "sourceRoot": str(root),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "allFileCount": len(all_files),
        "candidateDocumentCount": len(candidates),
        "extensionCounts": dict(ext_counter.most_common()),
        "statusCounts": dict(status_counter.most_common()),
        "outputs": {
            "reviewJson": "pte-knowledge.review.json",
            "sunnyDraftJson": "pte-knowledge.generated.json",
            "manifestCsv": "manifest.csv",
            "summaryJson": "summary.json",
        },
        "notes": [
            "This is a draft generated for review; it has not been merged into the live Sunny knowledge base.",
            "Video/audio files are indexed only in extension counts for v1 and are not transcribed.",
            "Large documents above max_extract_mb are indexed but not text-extracted.",
        ],
    }

    summary = sanitize_for_json(summary)
    review_entries = sanitize_for_json(review_entries)
    sunny_entries = sanitize_for_json(sunny_entries)
    manifest_rows = sanitize_for_json(manifest_rows)

    (out_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "pte-knowledge.review.json").write_text(json.dumps(review_entries, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "pte-knowledge.generated.json").write_text(json.dumps(sunny_entries, ensure_ascii=False, indent=2), encoding="utf-8")

    with (out_dir / "manifest.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["path", "relative_path", "extension", "size_mb", "status", "priority", "tags"])
        writer.writeheader()
        writer.writerows(manifest_rows)

    readme = [
        "# PTE Knowledge Draft Run",
        "",
        f"- Run ID: `{run_id}`",
        f"- Source root: `{root}`",
        f"- Documents processed: `{len(candidates)}`",
        f"- Outputs generated in: `{out_dir}`",
        "",
        "## Files",
        "",
        "- `summary.json`: run summary and counts",
        "- `manifest.csv`: every document candidate and extraction status",
        "- `pte-knowledge.review.json`: rich review file with previews and source paths",
        "- `pte-knowledge.generated.json`: Sunny-compatible draft entries, still needs human review",
        "",
        "## Next Step",
        "",
        "Open `pte-knowledge.review.json`, choose the entries worth keeping, then merge reviewed answers into `data/pte-knowledge.json`.",
        "Do not publish third-party course text verbatim; convert it into SunPace-owned guidance and answer style first.",
    ]
    (out_dir / "README.md").write_text("\n".join(readme) + "\n", encoding="utf-8")

    return {"outDir": str(out_dir), "summary": summary}


def main() -> None:
    logging.getLogger("pypdf").setLevel(logging.ERROR)
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", default="/Volumes/PTE_Resources")
    parser.add_argument("--output-dir", default="knowledge_exports")
    parser.add_argument("--max-files", type=int, default=0, help="0 means all supported documents")
    parser.add_argument("--max-pdf-pages", type=int, default=12)
    parser.add_argument("--max-chars", type=int, default=8000)
    parser.add_argument("--max-extract-mb", type=float, default=80)
    args = parser.parse_args()

    result = build(args)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
